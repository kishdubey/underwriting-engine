#!/usr/bin/env python3
"""
Script to analyze the Excel files and compare their contents
"""
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import sys
import os

def analyze_excel_file(filepath, name):
    """Analyze an Excel file and print key information"""
    print("\n" + "="*60)
    print("ANALYZING: " + name)
    print("File: " + filepath)
    print("="*60)
    
    if not os.path.exists(filepath):
        print("ERROR: File does not exist: " + filepath)
        return None
    
    try:
        wb = load_workbook(filepath, data_only=True)
        print("Number of sheets: " + str(len(wb.sheetnames)))
        print("Sheets: " + str(wb.sheetnames))
        
        # Check each sheet
        for sheet_name in wb.sheetnames:
            print("\n--- Sheet: " + sheet_name + " ---")
            ws = wb[sheet_name]
            
            # Read first few rows and columns
            max_row = min(30, ws.max_row)  # Read first 30 rows max
            max_col = min(15, ws.max_column)  # Read first 15 columns max
            
            print("Dimensions: " + str(ws.max_row) + " rows x " + str(ws.max_column) + " columns")
            
            # Print first few rows of data
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        value = str(cell.value)
                        # Truncate long values for readability
                        if len(value) > 25:
                            value = value[:22] + "..."
                        row_data.append(value)
                    else:
                        row_data.append("")
                
                # Only print if there's meaningful data in the row
                if any(row_data):
                    print("Row " + str(row) + ": " + str(row_data))
        
        wb.close()
        return True
        
    except Exception as e:
        print("ERROR reading " + filepath + ": " + str(e))
        return None

def main():
    print("Excel File Analysis Tool")
    print("Analysis Date: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    # Paths to the Excel files
    generated_file = "/Users/kishdubey/Documents/underwriting_mvp/underwriting_120_Valleywood_Drive_2025-10-15 (1).xlsx"
    analyst_file = "/Users/kishdubey/Documents/underwriting_mvp/120 Valleywood Markham.xlsx"
    
    # Analyze both files
    generated_data = analyze_excel_file(generated_file, "GENERATED REPORT")
    analyst_data = analyze_excel_file(analyst_file, "ANALYST EXCEL")
    
    # Compare key metrics if both files were analyzed successfully
    if generated_data and analyst_data:
        print("\n" + "="*60)
        print("COMPARISON SUMMARY")
        print("="*60)
        print("Comparison would be performed here if data extraction was implemented.")
    
    print("\nAnalysis completed at " + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

if __name__ == "__main__":
    main()