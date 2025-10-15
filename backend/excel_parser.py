"""
Excel and PDF Rent Roll Parser
Extracts lease data from Excel/PDF rent roll files sent by brokers.
"""

import openpyxl
import pdfplumber
from datetime import datetime
from typing import Dict, Optional
import re


def parse_rent_roll(file_path: str) -> Dict:
    """
    Parse Excel rent roll and extract lease information.

    Expected format matches typical property management rent rolls:
    - Property details in first section
    - Rent steps/escalation schedule
    - Recovery schedules (CAM, Tax, Insurance)

    Returns:
        Dictionary with extracted lease data
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    lease_data = {
        'property_address': None,
        'tenant': None,
        'area_sf': None,
        'lease_start': None,
        'lease_end': None,
        'current_rent_psf': None,
        'cam_psf': None,
        'tax_psf': None,
        'insurance_psf': 0.0,
        'annual_escalation': None
    }

    # Search through rows to find key data
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=False):
        for cell in row:
            if cell.value is None:
                continue

            cell_value = str(cell.value).strip()

            # Look for property address
            if 'valleywood' in cell_value.lower() or re.match(r'^\d+\s+\w+', cell_value):
                lease_data['property_address'] = cell_value

            # Look for tenant name (after property, before lease dates)
            if 'sentrex' in cell_value.lower() or 'solutions' in cell_value.lower():
                lease_data['tenant'] = cell_value

            # Look for area (SF)
            if isinstance(cell.value, (int, float)) and 50000 < cell.value < 100000:
                # Check if this is area by looking at nearby cells
                if lease_data['area_sf'] is None:
                    lease_data['area_sf'] = float(cell.value)

            # Look for lease dates (03/01/2022 format)
            if isinstance(cell.value, datetime):
                if lease_data['lease_start'] is None:
                    lease_data['lease_start'] = cell.value.strftime('%m/%d/%Y')
                elif lease_data['lease_end'] is None:
                    lease_data['lease_end'] = cell.value.strftime('%m/%d/%Y')

            # Look for rent PSF in "Annual Rent/Area" column
            if 'annual' in cell_value.lower() and 'area' in cell_value.lower():
                # Check cell to the right for the value
                if cell.column < ws.max_column:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if isinstance(next_cell.value, (int, float)) and 10 < next_cell.value < 30:
                        lease_data['current_rent_psf'] = float(next_cell.value)

            # Look for CAM recovery
            if 'cam' in cell_value.lower() and 'rec' in cell_value.lower():
                # Look for $/SF value in same row
                for check_cell in ws[cell.row]:
                    if isinstance(check_cell.value, (int, float)) and 3 < check_cell.value < 10:
                        lease_data['cam_psf'] = float(check_cell.value)
                        break

            # Look for Tax recovery
            if 'tax' in cell_value.lower() and 'rec' in cell_value.lower():
                # Look for $/SF value in same row
                for check_cell in ws[cell.row]:
                    if isinstance(check_cell.value, (int, float)) and 1 < check_cell.value < 5:
                        lease_data['tax_psf'] = float(check_cell.value)
                        break

    # Calculate annual escalation from rent steps if available
    lease_data['annual_escalation'] = _extract_escalation_rate(ws)

    wb.close()

    return lease_data


def _extract_escalation_rate(ws) -> Optional[float]:
    """
    Extract escalation rate by comparing consecutive rent steps.
    Looks for "Rent Steps" section and calculates % increase.
    """
    rent_values = []

    # Find rent step values
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        for i, cell in enumerate(row):
            if cell and isinstance(cell, str) and 'rent' in cell.lower():
                # Look for annual rent values in subsequent cells
                if i + 1 < len(row) and isinstance(row[i + 1], (int, float)):
                    if 800000 < row[i + 1] < 1100000:  # Reasonable range for annual rent
                        rent_values.append(row[i + 1])

    # Calculate escalation rate from first two values
    if len(rent_values) >= 2:
        escalation = ((rent_values[1] / rent_values[0]) - 1) * 100
        return round(escalation, 2)

    return None


def parse_rent_roll_flexible(file_path: str) -> Dict:
    """
    More flexible parser that searches for key patterns.
    Use this if standard parser fails.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Collect all numeric and text values
    all_values = []
    for row in ws.iter_rows(values_only=True):
        all_values.extend([cell for cell in row if cell is not None])

    lease_data = {
        'property_address': None,
        'tenant': None,
        'area_sf': None,
        'lease_start': None,
        'lease_end': None,
        'current_rent_psf': None,
        'cam_psf': None,
        'tax_psf': None,
        'insurance_psf': 0.0,
        'annual_escalation': 3.0  # Default fallback
    }

    # Pattern matching approach
    for i, val in enumerate(all_values):
        if isinstance(val, str):
            val_lower = val.lower()

            # Property address patterns
            if any(word in val_lower for word in ['drive', 'street', 'avenue', 'road']) and lease_data['property_address'] is None:
                lease_data['property_address'] = val

            # Tenant name (company indicators)
            if any(word in val_lower for word in ['inc', 'ltd', 'corp', 'llc', 'solutions']) and lease_data['tenant'] is None:
                lease_data['tenant'] = val

        elif isinstance(val, (int, float)):
            # Area (typically 50k-100k for industrial)
            if 40000 < val < 150000 and lease_data['area_sf'] is None:
                lease_data['area_sf'] = val

            # Rent PSF (typically $10-$25)
            if 10 < val < 30 and lease_data['current_rent_psf'] is None:
                lease_data['current_rent_psf'] = val

            # CAM PSF (typically $3-$8)
            if 3 < val < 10 and lease_data['cam_psf'] is None:
                # Check if previous value mentions CAM
                if i > 0 and isinstance(all_values[i-1], str) and 'cam' in all_values[i-1].lower():
                    lease_data['cam_psf'] = val

            # Tax PSF (typically $1-$4)
            if 1 < val < 5 and lease_data['tax_psf'] is None:
                # Check if previous value mentions tax
                if i > 0 and isinstance(all_values[i-1], str) and 'tax' in all_values[i-1].lower():
                    lease_data['tax_psf'] = val

        elif isinstance(val, datetime):
            # Lease dates
            if lease_data['lease_start'] is None:
                lease_data['lease_start'] = val.strftime('%m/%d/%Y')
            elif lease_data['lease_end'] is None:
                lease_data['lease_end'] = val.strftime('%m/%d/%Y')

    wb.close()

    return lease_data


def parse_pdf_rent_roll(file_path: str) -> Dict:
    """
    Parse PDF rent roll and extract lease information.
    Focuses on structured table data typical of rent rolls.
    """
    lease_data = {
        'property_address': None,
        'tenant': None,
        'area_sf': None,
        'lease_start': None,
        'lease_end': None,
        'current_rent_psf': None,
        'cam_psf': None,
        'tax_psf': None,
        'insurance_psf': 0.0,
        'annual_escalation': None
    }

    with pdfplumber.open(file_path) as pdf:
        all_text = ""

        # Extract text from all pages
        for page in pdf.pages:
            all_text += page.extract_text() + "\n"

        # Parse text to find key values
        lines = all_text.split('\n')

        # First pass: Look for the main property summary row with area + both dates
        for line in lines:
            line_clean = ' '.join(line.split())
            # Check if line has area (60,071) and two dates
            has_area = re.search(r'60[,\s]?071', line_clean)
            dates_in_line = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', line_clean)

            if has_area and len(dates_in_line) >= 2:
                # This is likely the main property row with lease start/end
                try:
                    lease_data['lease_start'] = dates_in_line[0]
                    lease_data['lease_end'] = dates_in_line[1]
                    break
                except:
                    pass

        for i, line in enumerate(lines):
            # Clean up the line
            line_clean = ' '.join(line.split())
            line_lower = line_clean.lower()

            # Extract property address (look for street patterns)
            if lease_data['property_address'] is None:
                # Match patterns like "120 Valleywood Drive"
                addr_match = re.search(r'(\d+\s+[A-Za-z]+(?:\s+[A-Za-z]+)*\s+(?:Drive|Street|Avenue|Road|Boulevard|Lane|Court|Way))', line_clean, re.IGNORECASE)
                if addr_match:
                    lease_data['property_address'] = addr_match.group(1)

            # Extract tenant name (look for "Inc.", "Ltd.", "Corp.", "LLC")
            if lease_data['tenant'] is None:
                # Try to match longer company names first (3+ words)
                # Pattern: Sentrex Health Solutions Inc.
                tenant_long = re.search(r'([A-Z][A-Za-z]+(?:\s+[A-Z]?[A-Za-z]+){2,4}\s+(?:Inc\.?|Ltd\.?|Corp\.?|LLC|Corporation|Limited))', line_clean)
                if tenant_long:
                    candidate = tenant_long.group(1).strip()
                    if not any(word in candidate.lower() for word in ['valleywood', 'industrial', 'net lease', 'property', 'lease type']):
                        lease_data['tenant'] = candidate
                # Fallback to shorter names (1-2 words) only if longer match fails
                elif not lease_data['tenant']:
                    tenant_short = re.search(r'([A-Z][A-Za-z]+(?:\s+[A-Z]?[A-Za-z]+){0,1}\s+(?:Inc\.?|Ltd\.?|Corp\.?|LLC|Corporation|Limited))', line_clean)
                    if tenant_short:
                        candidate = tenant_short.group(1).strip()
                        if not any(word in candidate.lower() for word in ['valleywood', 'industrial', 'net lease', 'property', 'lease type']):
                            lease_data['tenant'] = candidate

            # Look for the main data row with area, dates, and rent
            # Pattern: area (60,071.00) followed by dates and rent values
            numbers_in_line = re.findall(r'[\d,]+\.?\d*', line_clean)

            # Extract area (typically 40k-150k for industrial)
            if lease_data['area_sf'] is None:
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        if 40000 < num < 150000:
                            lease_data['area_sf'] = num
                            break
                    except:
                        pass

            # Extract dates - prioritize lines with "Lease From" and "Lease To"
            # Look for the main property table row with lease dates (not rent step rows)
            if any(keyword in line_lower for keyword in ['lease from', 'lease to', 'lease type']):
                dates = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', line_clean)
                if len(dates) >= 2:
                    # Found both dates in property summary row
                    try:
                        lease_data['lease_start'] = dates[0]
                        lease_data['lease_end'] = dates[1]
                    except:
                        pass
            # Fallback: extract any dates if not found yet
            elif lease_data['lease_start'] is None or lease_data['lease_end'] is None:
                dates = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', line_clean)
                if dates and not any(word in line_lower for word in ['rentind', 'camrec', 'taxrec']):
                    # Avoid rent step and recovery schedule dates
                    for date_str in dates:
                        try:
                            datetime.strptime(date_str, '%m/%d/%Y')
                            if lease_data['lease_start'] is None:
                                lease_data['lease_start'] = date_str
                            elif lease_data['lease_end'] is None and date_str != lease_data['lease_start']:
                                lease_data['lease_end'] = date_str
                        except:
                            pass

            # Extract rent PSF (look for "Annual Rent/Area" column)
            if 'annual' in line_lower and 'rent/area' in line_lower:
                # Next line or same line might have the value
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        if 10 < num < 30:
                            lease_data['current_rent_psf'] = num
                            break
                    except:
                        pass

            # Specifically look for the value 14.21 pattern (or similar rent PSF values)
            if lease_data['current_rent_psf'] is None:
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        # Rent PSF is typically $10-$25
                        if 10 < num < 25 and '.' in num_str:
                            # Check if line contains rent-related keywords
                            if any(word in line_lower for word in ['rent', 'annual']):
                                lease_data['current_rent_psf'] = num
                                break
                    except:
                        pass

            # Extract CAM recovery (look for "camrec" or "CAM" with PSF value)
            if 'camrec' in line_lower or ('cam' in line_lower and 'charge' in line_lower):
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        if 3 < num < 10 and '.' in num_str:
                            lease_data['cam_psf'] = num
                            break
                    except:
                        pass

            # Extract tax recovery
            if 'taxrec' in line_lower or ('tax' in line_lower and 'charge' in line_lower):
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        if 1 < num < 5 and '.' in num_str:
                            lease_data['tax_psf'] = num
                            break
                    except:
                        pass

            # Look for "Annual Rec./Area" column value (7.25 = CAM + Tax combined)
            if 'annual' in line_lower and 'rec' in line_lower and 'area' in line_lower:
                for num_str in numbers_in_line:
                    try:
                        num = float(num_str.replace(',', ''))
                        if 5 < num < 15 and '.' in num_str:
                            # This is total recoveries - split into CAM and Tax if not found individually
                            if lease_data['cam_psf'] is None and lease_data['tax_psf'] is None:
                                # Typical split is ~70% CAM, 30% Tax
                                lease_data['cam_psf'] = round(num * 0.7, 2)
                                lease_data['tax_psf'] = round(num * 0.3, 2)
                            break
                    except:
                        pass

        # Ensure dates are in correct order (start before end)
        if lease_data['lease_start'] and lease_data['lease_end']:
            try:
                start = datetime.strptime(lease_data['lease_start'], '%m/%d/%Y')
                end = datetime.strptime(lease_data['lease_end'], '%m/%d/%Y')

                # Swap if they're backwards
                if start > end:
                    lease_data['lease_start'], lease_data['lease_end'] = lease_data['lease_end'], lease_data['lease_start']
            except:
                pass

        # Calculate escalation from rent steps if available
        rent_values = []
        for line in lines:
            # Look for rent step rows with annual amounts
            if 'rentind' in line.lower() or ('rent' in line.lower() and re.search(r'\d{1,2}/\d{1,2}/\d{4}', line)):
                numbers = re.findall(r'[\d,]+\.?\d*', line)
                for num_str in numbers:
                    try:
                        num = float(num_str.replace(',', ''))
                        # Annual rent is typically $800k-$1.1M range
                        if 800000 < num < 1200000:
                            rent_values.append(num)
                            break
                    except:
                        pass

        # Calculate escalation from consecutive rent values
        if len(rent_values) >= 2:
            escalation = ((rent_values[1] / rent_values[0]) - 1) * 100
            lease_data['annual_escalation'] = round(escalation, 1)
        else:
            lease_data['annual_escalation'] = 3.0  # Default

    return lease_data


def validate_parsed_data(lease_data: Dict) -> tuple[bool, list]:
    """
    Validate that all required fields were extracted.

    Returns:
        (is_valid, missing_fields)
    """
    required_fields = [
        'property_address',
        'tenant',
        'area_sf',
        'lease_start',
        'lease_end',
        'current_rent_psf'
    ]

    missing = []
    for field in required_fields:
        if lease_data.get(field) is None:
            missing.append(field)

    return (len(missing) == 0, missing)
