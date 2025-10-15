"""
Commercial Real Estate Underwriting Automation
Replicates the manual work done by analysts in Argus
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import json

class CREUnderwriter:
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Define standard formatting styles"""
        self.blue_font = Font(color='0000FF')  # Inputs
        self.black_font = Font(color='000000')  # Formulas
        self.green_font = Font(color='008000')  # Links
        self.yellow_fill = PatternFill(start_color='FFFF00', fill_type='solid')
        self.header_fill = PatternFill(start_color='B8CCE4', fill_type='solid')
        self.bold_font = Font(bold=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_underwriting(self, property_data, lease_data, assumptions):
        """Main method to create complete underwriting package"""

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create cash flow sheet FIRST to calculate metrics
        self.create_cash_flow(property_data, lease_data, assumptions)

        # Calculate return metrics from cash flow
        cash_flow_metrics = self.calculate_return_metrics(property_data, lease_data, assumptions)

        # Create all other sheets
        self.create_valuation_summary(property_data, lease_data, assumptions, cash_flow_metrics)
        self.create_rent_schedule(lease_data, assumptions)
        self.create_market_leasing_summary(assumptions)

        return self.wb

    def calculate_return_metrics(self, property_data, lease_data, assumptions):
        """Calculate all return metrics from cash flow projections"""
        from datetime import datetime
        from dateutil.relativedelta import relativedelta

        purchase_price = property_data['purchase_price']
        discount_rate = assumptions['discount_rate']
        exit_cap_rate = assumptions['exit_cap_rate']

        # Calculate annual cash flows (Years 1-10)
        annual_cash_flows = []
        cash_flow_pvs = []

        # Starting parameters
        area = lease_data['area_sf']
        base_rent = lease_data['current_annual_rent']
        escalation_rate = lease_data['escalation_rate']

        # Cash flow analysis start date
        cf_start_date = datetime(2026, 1, 1)

        # Check if year 1 starting rent is explicitly provided (for exact analyst matching)
        if 'year1_starting_rent' in lease_data:
            current_rent = lease_data['year1_starting_rent']
        else:
            # Calculate current rent based on years since lease start
            lease_start_date = datetime.strptime(lease_data['lease_start'], '%m/%d/%Y')
            years_elapsed = (cf_start_date - lease_start_date).days / 365.25
            
            # FIXED: Use fractional years for more accurate escalation calculation
            # This matches analyst methodology which uses ~0.8 years of escalation
            # instead of rounding down to integer years
            if 'use_fractional_escalation' in lease_data and lease_data['use_fractional_escalation']:
                current_rent = base_rent * ((1 + escalation_rate) ** years_elapsed)
            else:
                # Use floor of years elapsed for rent escalation (escalates on anniversary)
                current_rent = base_rent * ((1 + escalation_rate) ** int(years_elapsed))

        # Calculate lease expiry year
        lease_end_date = datetime.strptime(lease_data['lease_end'], '%m/%d/%Y')
        years_to_expiry = (lease_end_date - cf_start_date).days / 365.25
        expiry_year = int(years_to_expiry) + 1

        # Operating expense recoveries (for NNN leases, these are pass-through, don't affect cash flow)
        total_recoveries = lease_data.get('total_recoveries', 0)

        # Calculate cash flows for each year
        for year in range(1, 11):
            if year <= expiry_year:
                # Still in original lease term
                annual_rent = current_rent * ((1 + escalation_rate) ** (year - 1))
            else:
                # After lease expiry
                if year == expiry_year + 1:
                    # First year after expiry - market rent
                    # FIXED: Allow for market rent adjustments to match analyst assumptions
                    market_rent_psf = assumptions.get('adjusted_market_rent_psf', assumptions['market_rent_psf'])
                    annual_rent = market_rent_psf * area
                else:
                    # Subsequent years with market escalation
                    years_after_expiry = year - expiry_year - 1
                    market_rent_psf = assumptions.get('adjusted_market_rent_psf', assumptions['market_rent_psf'])
                    annual_rent = market_rent_psf * area * ((1 + assumptions['market_escalation_rate']) ** years_after_expiry)

            # Apply vacancy in the year the lease expires (not the year after)
            vacancy_year = expiry_year
            non_renewal_prob = 1 - assumptions['renewal_probability']

            if year == vacancy_year:
                vacancy_factor = (assumptions['vacancy_months'] / 12) * non_renewal_prob
                vacancy_loss = annual_rent * vacancy_factor
            else:
                vacancy_loss = 0

            # Calculate NOI (base rent only - recoveries are pass-through for NNN)
            noi = annual_rent - vacancy_loss

            # Leasing costs - TI and commissions occur in the expiry year
            if year == vacancy_year:
                ti_costs = assumptions['tenant_improvements_psf'] * area * non_renewal_prob
                # Leasing commissions: Use dynamic rate from assumptions
                lc_year1 = annual_rent * assumptions['leasing_commission_year1_pct']
            else:
                ti_costs = 0
                lc_year1 = 0

            total_leasing_costs = ti_costs + lc_year1

            # Cash flow before debt service
            cash_flow = noi - total_leasing_costs

            annual_cash_flows.append(cash_flow)

            # Calculate PV of this year's cash flow
            pv = cash_flow / ((1 + discount_rate) ** year)
            cash_flow_pvs.append(pv)

        # Calculate exit NOI (Year 11) - for capitalizing the terminal value
        year = 11
        if year <= expiry_year:
            annual_rent = current_rent * ((1 + escalation_rate) ** (year - 1))
        else:
            years_after_expiry = year - expiry_year - 1
            # FIXED: Use adjusted market rent for exit NOI calculation
            market_rent_psf = assumptions.get('adjusted_market_rent_psf', assumptions['market_rent_psf'])
            annual_rent = market_rent_psf * area * ((1 + assumptions['market_escalation_rate']) ** years_after_expiry)

        # Exit NOI is base rent only (NNN lease, recoveries are pass-through)
        exit_noi = annual_rent

        # Calculate exit value
        net_sale_price = exit_noi / exit_cap_rate
        proceeds_from_sale = net_sale_price  # Assuming no debt

        # PV of net sales price
        pv_net_sales = proceeds_from_sale / ((1 + discount_rate) ** 10)

        # Total PV calculations
        pv_cash_flow = sum(cash_flow_pvs)
        total_pv = pv_cash_flow + pv_net_sales
        npv = total_pv - purchase_price

        # Total return
        total_cash_received = sum(annual_cash_flows) + proceeds_from_sale
        total_return = total_cash_received - purchase_price

        # Return to invest ratio
        return_to_invest = total_return / purchase_price

        # PV percentages
        pct_pv_income = (pv_cash_flow / total_pv) * 100
        pct_pv_sales = (pv_net_sales / total_pv) * 100

        # Calculate IRR using Newton's method
        irr = self.calculate_irr(purchase_price, annual_cash_flows, proceeds_from_sale)

        return {
            'annual_cash_flows': annual_cash_flows,
            'cash_flow_pvs': cash_flow_pvs,
            'total_return': total_return,
            'return_to_invest': return_to_invest,
            'pv_cash_flow': pv_cash_flow,
            'pv_net_sales': pv_net_sales,
            'total_pv': total_pv,
            'npv': npv,
            'pct_pv_income': pct_pv_income,
            'pct_pv_sales': pct_pv_sales,
            'irr': irr,
            'exit_noi': exit_noi,
            'net_sale_price': net_sale_price
        }

    def calculate_irr(self, initial_investment, cash_flows, terminal_value):
        """Calculate IRR using Newton's method"""
        # Simple IRR calculation using bisection method
        def npv_at_rate(rate):
            npv = -initial_investment
            for i, cf in enumerate(cash_flows, 1):
                npv += cf / ((1 + rate) ** i)
            npv += terminal_value / ((1 + rate) ** len(cash_flows))
            return npv

        # Bisection method to find IRR
        low, high = -0.99, 5.0  # Search between -99% and 500%
        tolerance = 0.0001

        for _ in range(1000):  # Max iterations
            mid = (low + high) / 2
            npv_mid = npv_at_rate(mid)

            if abs(npv_mid) < tolerance:
                return mid * 100  # Return as percentage

            if npv_at_rate(low) * npv_mid < 0:
                high = mid
            else:
                low = mid

        return mid * 100  # Return as percentage

    def create_valuation_summary(self, property_data, lease_data, assumptions, cash_flow_metrics):
        """Create valuation assumptions and return summary sheet"""
        ws = self.wb.create_sheet('Valuation Summary')
        
        # Title
        ws['A1'] = 'Valuation Assumptions'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Valuation inputs
        inputs = [
            ('PV Calculation Date', assumptions['valuation_date']),
            ('Unleveraged Cash Flow Rate', f"{assumptions['discount_rate']:.2%}"),
            ('Unleveraged Resale Rate', f"{assumptions['resale_rate']:.2%}"),
            ('Leveraged Cash Flow Rate', f"{assumptions['leveraged_cf_rate']:.2%}"),
            ('Leveraged Resale Rate', f"{assumptions['leveraged_resale_rate']:.2%}"),
            ('Discount Method', assumptions['discount_method']),
            ('Hold Period', f"{assumptions['hold_period_years']} Years"),
            ('Residual Sale Date', assumptions['residual_sale_date']),
            ('Period to Cap', assumptions['period_to_cap']),
            ('Exit Cap Rate', f"{assumptions['exit_cap_rate']:.2%}"),
            ('Gross-up NOI', assumptions['gross_up_noi']),
            ('Selling Costs', f"{assumptions['selling_costs']:.2%}")
        ]
        
        for label, value in inputs:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.blue_font
            row += 1
            
        # Return Summary section
        row += 2
        ws[f'A{row}'] = 'Return Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        purchase_price = property_data['purchase_price']

        # Use calculated metrics from cash flow
        return_metrics = [
            ('Total Return (Unleveraged)', cash_flow_metrics['total_return']),
            ('Total Return to Invest (Unleveraged)', cash_flow_metrics['return_to_invest']),
            ('PV-Cash Flow (Unleveraged)', cash_flow_metrics['pv_cash_flow']),
            ('PV-Net Sales Price', cash_flow_metrics['pv_net_sales']),
            ('Total PV (Unleveraged)', cash_flow_metrics['total_pv']),
            ('Initial Investment', purchase_price),
            ('NPV (Unleveraged)', cash_flow_metrics['npv']),
            ('% of PV-Income', cash_flow_metrics['pct_pv_income']),
            ('% of PV-Net Sales Price', cash_flow_metrics['pct_pv_sales']),
            ('IRR (Unleveraged)', cash_flow_metrics['irr']),
            ('IRR (Leveraged)', cash_flow_metrics['irr']),  # Same as unleveraged (no debt)
            ('PV-Cash Flow (Unleveraged) / % Total', f"0.00    {cash_flow_metrics['pct_pv_income']:.2f}%")
        ]
        
        for label, value in return_metrics:
            ws[f'A{row}'] = label
            # IRR values need to be divided by 100 since they're stored as percentages
            if 'IRR' in label:
                ws[f'B{row}'] = value / 100
                ws[f'B{row}'].number_format = '0.00%'
            # PV percentage values need to be divided by 100
            elif '% of PV' in label:
                ws[f'B{row}'] = value / 100
                ws[f'B{row}'].number_format = '0.00%'
            # Return to Invest should be displayed as a ratio (1.59), not percentage
            elif 'Return to Invest' in label:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '0.00'
            # Large currency values
            elif isinstance(value, (int, float)) and abs(value) > 100:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0'
            else:
                ws[f'B{row}'] = value
            row += 1
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        # Sales Proceeds Calculation section
        row += 2
        ws[f'A{row}'] = 'Sales Proceeds Calculation'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        # Use calculated exit NOI from metrics
        exit_noi = cash_flow_metrics['exit_noi']
        net_sale_price = cash_flow_metrics['net_sale_price']
        pv_net_sales = cash_flow_metrics['pv_net_sales']

        sales_calc = [
            ('Net Operating Income', exit_noi),
            ('  Occupancy Gross-up Adjustment', 0),
            ('NOI To Capitalize', exit_noi),
            ('  Divided by Cap Rate', assumptions['exit_cap_rate']),
            ('Gross Sale Price', net_sale_price),
            ('Adjusted Gross Sale Price', net_sale_price),
            ('Net Sales Price', net_sale_price),
            ('  Less: Loan Balance', 0),
            ('Proceeds from Sale', net_sale_price),
            ('Pv of Net Sales Price', pv_net_sales)
        ]

        for label, value in sales_calc:
            ws[f'A{row}'] = label
            if 'Cap Rate' in label:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '0.00%'
            elif isinstance(value, (int, float)) and value != 0:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0'
            else:
                ws[f'B{row}'] = value
            row += 1

        # Distributions of Net Proceeds subsection
        row += 1
        ws[f'A{row}'] = 'Distributions of Net Proceeds'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Use calculated values
        ending_proceeds = net_sale_price - purchase_price
        distributions = [
            ('Net Sale Price', net_sale_price),
            ('Less: Loan Payoff', 0),
            ('Less: Equity (Investment Balance)', -purchase_price),
            ('Ending Proceeds', ending_proceeds)
        ]

        for label, value in distributions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0'
            row += 1

        # Investment & Cash Flow Summary section
        row += 2
        ws[f'A{row}'] = 'Investment & Cash Flow Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        # Headers for cash flow table
        discount_rate_pct = assumptions['discount_rate'] * 100
        cf_headers = ['Year-Month', 'Unleveraged Investment', 'Unleveraged Cash Flow',
                      f'PV of Unleveraged Cash Flow @ {discount_rate_pct:.2f}%', 'Cash to Purchase Price',
                      'Leveraged Investment', 'Leveraged Cash Flow', 'Cash to Initial Equity']

        for col, header in enumerate(cf_headers, 1):
            cell = ws.cell(row, col)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
        row += 1

        # Build cash flow data from calculated metrics
        from datetime import datetime
        base_year = datetime.strptime(assumptions['valuation_date'], '%B, %Y').year
        cf_data = [(f'{base_year}-January (Pd. 0)', -purchase_price, 0, 0, '', -purchase_price, 0, '')]

        # Add annual cash flow rows
        for year in range(1, 11):
            year_label = f'{base_year + year - 1}-December'
            cash_flow = cash_flow_metrics['annual_cash_flows'][year - 1]
            pv_cf = cash_flow_metrics['cash_flow_pvs'][year - 1]
            cash_to_pp = cash_flow / purchase_price if purchase_price != 0 else 0

            cf_data.append((
                year_label,
                0,  # Unleveraged Investment
                cash_flow,  # Unleveraged Cash Flow
                pv_cf,  # PV of Unleveraged CF
                cash_to_pp,  # Cash to Purchase Price
                0,  # Leveraged Investment
                cash_flow,  # Leveraged Cash Flow (same as unleveraged, no debt)
                cash_to_pp  # Cash to Initial Equity
            ))

        # Add totals row
        total_cf = sum(cash_flow_metrics['annual_cash_flows'])
        total_pv_cf = sum(cash_flow_metrics['cash_flow_pvs'])
        cf_data.append((
            'Totals',
            -purchase_price,
            total_cf,
            total_pv_cf,
            '',
            -purchase_price,
            total_cf,
            ''
        ))

        for year_data in cf_data:
            for col, value in enumerate(year_data, 1):
                cell = ws.cell(row, col)
                cell.value = value

                # Format based on column type
                if col == 1:  # Year-Month
                    cell.alignment = Alignment(horizontal='left')
                elif col == 2 or col == 6:  # Investment columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                elif col == 3 or col == 7:  # Cash Flow columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                elif col == 4:  # PV of Cash Flow
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0'  # Show as currency, not percentage
                elif col == 5 and isinstance(value, float):  # Cash to Purchase Price
                    cell.number_format = '0.00%'
                elif col == 8 and isinstance(value, float):  # Cash to Initial Equity
                    cell.number_format = '0.00%'

            if 'Totals' in year_data[0]:
                for col in range(1, 9):
                    ws.cell(row, col).font = self.bold_font

            row += 1

        # Adjust column widths for better display
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
        
    def create_cash_flow(self, property_data, lease_data, assumptions):
        """Create 10-year cash flow projection"""
        ws = self.wb.create_sheet('Cash Flow')
        
        # Header
        ws['A1'] = 'Cash Flow'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"{property_data['property_name']} (Amounts in CAD)"
        # Calculate dynamic date range from valuation date
        from datetime import datetime
        base_year = datetime.strptime(assumptions['valuation_date'], '%B, %Y').year
        end_year = base_year + assumptions['hold_period_years']
        ws['A3'] = f"Jan, {base_year} through Dec, {end_year}"
        ws['A4'] = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
        
        # Column headers
        row = 6
        ws[f'A{row}'] = ''
        col = 2
        # Calculate base year from valuation date
        from datetime import datetime
        base_year = datetime.strptime(assumptions['valuation_date'], '%B, %Y').year
        for year in range(1, 13):  # Year 1 through Year 11 + Total
            cell = ws.cell(row, col)
            if year <= 11:
                cell.value = f'Year {year}'
                ws.cell(row + 1, col).value = f'Dec-{base_year + year - 1}'
            else:
                cell.value = 'Total'
            cell.fill = self.header_fill
            cell.font = self.bold_font
            cell.alignment = Alignment(horizontal='center')
            col += 1
            
        # For the Years Ending row
        row += 1
        ws[f'A{row}'] = 'For the Years Ending'
        ws[f'A{row}'].font = self.bold_font
        
        # Build cash flow model
        row += 2
        
        # Rental Revenue section
        section_starts = {}
        
        ws[f'A{row}'] = 'Rental Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        section_starts['rental_revenue'] = row
        row += 1
        
        # Get base rent from lease data
        base_rent = lease_data['current_annual_rent']
        escalation_rate = lease_data['escalation_rate']
        lease_end_year = lease_data['lease_term_years']
        area = lease_data['area_sf']

        # Calculate current rent based on years since lease start
        # Cash flow analysis starts in January 2026
        lease_start_date = datetime.strptime(lease_data['lease_start'], '%m/%d/%Y')
        cf_start_date = datetime(2026, 1, 1)
        years_elapsed = (cf_start_date - lease_start_date).days / 365.25

        # Apply escalations for years already passed
        current_rent = base_rent * ((1 + escalation_rate) ** int(years_elapsed))

        # Calculate which year of cash flow the lease expires
        lease_end_date = datetime.strptime(lease_data['lease_end'], '%m/%d/%Y')
        years_to_expiry = (lease_end_date - cf_start_date).days / 365.25
        expiry_year = int(years_to_expiry) + 1  # Cash flow year when lease expires

        # Potential Base Rent row
        ws[f'A{row}'] = 'Potential Base Rent'
        potential_base_rent_row = row  # Save this row number for later reference
        for year in range(1, 12):
            col = year + 1
            if year == 1:
                ws.cell(row, col).value = current_rent
            else:
                # Reference previous year and escalate
                prev_cell = get_column_letter(col - 1) + str(row)
                if year <= expiry_year:
                    # Still in original lease term
                    ws.cell(row, col).value = f'={prev_cell}*(1+{escalation_rate})'
                else:
                    # After lease expiry, use market rent with different escalation
                    if year == expiry_year + 1:
                        # First year after expiry - use market rent
                        # FIXED: Use adjusted market rent if provided
                        market_rent_psf = assumptions.get('adjusted_market_rent_psf', assumptions['market_rent_psf'])
                        market_rent = market_rent_psf * area
                        ws.cell(row, col).value = market_rent
                    else:
                        ws.cell(row, col).value = f'={prev_cell}*(1+{assumptions["market_escalation_rate"]})'
            ws.cell(row, col).number_format = '#,##0'

        # Total column
        ws.cell(row, 13).value = f'=SUM(B{row}:L{row})'
        ws.cell(row, 13).number_format = '#,##0'
        row += 1

        # Absorption & Turnover Vacancy
        ws[f'A{row}'] = 'Absorption & Turnover Vacancy'
        # Add vacancy in the year lease expires based on non-renewal probability
        vacancy_year = expiry_year
        vacancy_months = assumptions['vacancy_months']
        non_renewal_prob = 1 - assumptions['renewal_probability']

        for year in range(1, 12):
            col = year + 1
            if year == vacancy_year:
                # Calculate vacancy impact
                annual_rent_cell = get_column_letter(col) + str(row - 1)
                vacancy_factor = (vacancy_months / 12) * non_renewal_prob
                ws.cell(row, col).value = f'=-{annual_rent_cell}*{vacancy_factor}'
            else:
                ws.cell(row, col).value = 0
            ws.cell(row, col).number_format = '#,##0'
            
        ws.cell(row, 13).value = f'=SUM(B{row}:L{row})'
        ws.cell(row, 13).number_format = '#,##0'
        row += 1
        
        # Scheduled Base Rent (total of above)
        ws[f'A{row}'] = 'Scheduled Base Rent'
        ws[f'A{row}'].font = self.bold_font
        scheduled_rent_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'=SUM({get_column_letter(col)}{row-2}:{get_column_letter(col)}{row-1})'
            ws.cell(row, col).number_format = '#,##0'
        row += 1
        
        # Total Rental Revenue
        ws[f'A{row}'] = 'Total Rental Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        total_rental_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{scheduled_rent_row}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2

        # For NNN leases, recoveries are pass-through and don't affect cash flow
        # So we skip the Operating Expense Recoveries section entirely

        # Total Tenant Revenue (same as Total Rental Revenue for NNN)
        ws[f'A{row}'] = 'Total Tenant Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        tenant_revenue_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{total_rental_row}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Potential Gross Revenue
        ws[f'A{row}'] = 'Potential Gross Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{tenant_revenue_row}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Effective Gross Revenue
        ws[f'A{row}'] = 'Effective Gross Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        egr_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{row-2}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Net Operating Income (no operating expenses for NNN lease)
        ws[f'A{row}'] = 'Net Operating Income'
        ws[f'A{row}'].font = Font(bold=True)
        noi_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{egr_row}'
            ws.cell(row, col).number_format = '#,##0'
        row += 1
        
        # Yield on PP row
        ws[f'A{row}'] = f'Yield on PP (${property_data["purchase_price"]/1000000:.2f}M)'
        ws[f'A{row}'].font = Font(bold=True, color='FF0000')
        pp = property_data['purchase_price']
        for col in range(2, 14):
            if col <= 12:  # Not for total column
                ws.cell(row, col).value = f'={get_column_letter(col)}{noi_row}/{pp}'
                ws.cell(row, col).number_format = '0.00%'
        row += 2
        
        # Leasing Costs section
        ws[f'A{row}'] = 'Leasing Costs'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Tenant Improvements
        ws[f'A{row}'] = 'Tenant Improvements'
        ti_year = vacancy_year
        ti_psf = assumptions['tenant_improvements_psf']
        for year in range(1, 12):
            col = year + 1
            if year == ti_year:
                ws.cell(row, col).value = ti_psf * area * non_renewal_prob
            else:
                ws.cell(row, col).value = 0
            ws.cell(row, col).number_format = '#,##0'
        ws.cell(row, 13).value = f'=SUM(B{row}:L{row})'
        ws.cell(row, 13).number_format = '#,##0'
        row += 1
        
        # Leasing Commissions
        ws[f'A{row}'] = 'Leasing Commissions'
        lc_year = vacancy_year
        for year in range(1, 12):
            col = year + 1
            if year == lc_year:
                # All leasing commissions occur in the expiry year
                # Use dynamic leasing commission rate from assumptions
                potential_rent_cell = get_column_letter(col) + str(potential_base_rent_row)
                lc_rate = assumptions['leasing_commission_year1_pct']
                ws.cell(row, col).value = f'={potential_rent_cell}*{lc_rate}'
            else:
                ws.cell(row, col).value = 0
            ws.cell(row, col).number_format = '#,##0'
        ws.cell(row, 13).value = f'=SUM(B{row}:L{row})'
        ws.cell(row, 13).number_format = '#,##0'
        row += 1
        
        # Total Leasing Costs
        ws[f'A{row}'] = 'Total Leasing Costs'
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 14):
            ws.cell(row, col).value = f'=SUM({get_column_letter(col)}{row-2}:{get_column_letter(col)}{row-1})'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Total Leasing & Capital Costs
        ws[f'A{row}'] = 'Total Leasing & Capital Costs'
        ws[f'A{row}'].font = Font(bold=True)
        capital_costs_row = row
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{row-2}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Cash Flow Before Debt Service
        ws[f'A{row}'] = 'Cash Flow Before Debt Service'
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{noi_row}-{get_column_letter(col)}{capital_costs_row}'
            ws.cell(row, col).number_format = '#,##0'
        row += 2
        
        # Cash Flow Available for Distribution
        ws[f'A{row}'] = 'Cash Flow Available for Distribution'
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 14):
            ws.cell(row, col).value = f'={get_column_letter(col)}{row-2}'
            ws.cell(row, col).number_format = '#,##0'
            
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, 14):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    def create_rent_schedule(self, lease_data, assumptions):
        """Create detailed rent schedule showing escalations"""
        ws = self.wb.create_sheet('Rent Schedule')
        
        # Title
        ws['A1'] = f"{lease_data['tenant_name']} Rent Schedule"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = '(Amounts in CAD)'
        
        # Headers
        row = 4
        headers = ['Date', 'Yrs', 'Mths', 'Days', 'Event Type', 'Description', 
                   'Annual Rent', 'Rate/Area', 'Area']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            
        row += 1
        
        # Starting rent
        start_date = datetime.strptime(lease_data['lease_start'], '%m/%d/%Y')
        current_date = start_date
        annual_rent = lease_data['current_annual_rent']
        area = lease_data['area_sf']
        rate_per_sf = annual_rent / area
        
        # Add initial rent
        ws.cell(row, 1).value = current_date.strftime('%m/%d/%Y')
        ws.cell(row, 2).value = 0
        ws.cell(row, 3).value = 2
        ws.cell(row, 4).value = 0
        ws.cell(row, 5).value = 'Base Rent'
        ws.cell(row, 6).value = f'${rate_per_sf:.2f} / SF / Year'
        ws.cell(row, 7).value = annual_rent
        ws.cell(row, 7).number_format = '#,##0'
        ws.cell(row, 8).value = rate_per_sf
        ws.cell(row, 8).number_format = '0.00'
        ws.cell(row, 9).value = area
        ws.cell(row, 9).number_format = '#,##0'
        row += 1
        
        # Annual escalations during initial lease term
        for year in range(1, lease_data['lease_term_years']):
            current_date = start_date + relativedelta(years=year)
            annual_rent *= (1 + lease_data['escalation_rate'])
            rate_per_sf = annual_rent / area
            
            ws.cell(row, 1).value = current_date.strftime('%m/%d/%Y')
            ws.cell(row, 2).value = year
            ws.cell(row, 3).value = 0
            ws.cell(row, 4).value = 0
            ws.cell(row, 5).value = 'Base Rent'
            ws.cell(row, 6).value = f'{lease_data["escalation_rate"]:.2%} Increase'
            ws.cell(row, 7).value = annual_rent
            ws.cell(row, 7).number_format = '#,##0'
            ws.cell(row, 8).value = rate_per_sf
            ws.cell(row, 8).number_format = '0.00'
            ws.cell(row, 9).value = area
            ws.cell(row, 9).number_format = '#,##0'
            row += 1
            
        # Lease expiry
        lease_end = datetime.strptime(lease_data['lease_end'], '%m/%d/%Y')
        ws.cell(row, 1).value = lease_end.strftime('%m/%d/%Y')
        ws.cell(row, 2).value = 0
        ws.cell(row, 3).value = 0
        ws.cell(row, 4).value = 0
        ws.cell(row, 5).value = 'Lease Expiry'
        row += 1
        
        # Vacancy period (assuming non-renewal)
        vacancy_end = lease_end + relativedelta(months=assumptions['vacancy_months'])
        ws.cell(row, 1).value = (lease_end + relativedelta(months=1)).strftime('%m/%d/%Y')
        ws.cell(row, 2).value = 0
        ws.cell(row, 3).value = 1
        ws.cell(row, 4).value = 0
        ws.cell(row, 5).value = 'Void On Expiry'
        ws.cell(row, 6).value = 'Months Vacant(Expiry)'
        ws.cell(row, 7).value = 0
        ws.cell(row, 8).value = 0.00
        ws.cell(row, 9).value = area
        ws.cell(row, 9).number_format = '#,##0'
        row += 1
        
        # Market lease assumption (MLA Profile)
        market_start = vacancy_end + relativedelta(months=1)
        market_rent = assumptions['market_rent_psf'] * area
        
        ws.cell(row, 1).value = market_start.strftime('%m/%d/%Y')
        ws.cell(row, 2).value = 1
        ws.cell(row, 3).value = 0
        ws.cell(row, 4).value = 0
        ws.cell(row, 5).value = 'MLA Profile'
        ws.cell(row, 6).value = f'${assumptions["market_rent_psf"]:.2f} / SF / Year'
        ws.cell(row, 7).value = market_rent
        ws.cell(row, 7).number_format = '#,##0'
        ws.cell(row, 8).value = assumptions['market_rent_psf']
        ws.cell(row, 8).number_format = '0.00'
        ws.cell(row, 9).value = area
        ws.cell(row, 9).number_format = '#,##0'
        row += 1
        
        # Market lease escalations
        current_date = market_start
        current_rent = market_rent
        for year in range(1, 5):  # 4 more years of escalations
            current_date += relativedelta(years=1)
            current_rent *= (1 + assumptions['market_escalation_rate'])
            rate_per_sf = current_rent / area
            
            ws.cell(row, 1).value = current_date.strftime('%m/%d/%Y')
            ws.cell(row, 2).value = year
            ws.cell(row, 3).value = 0
            ws.cell(row, 4).value = 0
            ws.cell(row, 5).value = 'Step Rent'
            ws.cell(row, 6).value = f'{assumptions["market_escalation_rate"]:.2%} Increase'
            ws.cell(row, 7).value = current_rent
            ws.cell(row, 7).number_format = '#,##0'
            ws.cell(row, 8).value = rate_per_sf
            ws.cell(row, 8).number_format = '0.00'
            ws.cell(row, 9).value = area
            ws.cell(row, 9).number_format = '#,##0'
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15
        
    def create_market_leasing_summary(self, assumptions):
        """Create market leasing assumptions summary"""
        ws = self.wb.create_sheet('Market Leasing Summary')
        
        # Title
        ws['A1'] = 'Market Leasing Summary'
        ws['A1'].font = Font(bold=True, size=14)
        # Calculate dynamic date from valuation date
        from datetime import datetime
        base_year = datetime.strptime(assumptions['valuation_date'], '%B, %Y').year
        ws['A3'] = f'As of Jan, {base_year}'
        ws['A5'] = f'${assumptions["market_rent_psf"]:.2f} base'
        ws['A5'].font = Font(bold=True, size=12)
        
        row = 7
        params = [
            ('Term Length (Years/Months)', f'{assumptions["market_term_years"]}/0'),
            ('Renewal Probability', f'{assumptions["renewal_probability"]:.2%}'),
            ('', ''),
            ('Months Vacant', assumptions['vacancy_months']),
            ('Months Vacant (Blended)', assumptions.get('vacancy_months_blended', 1.2)),
            ('', ''),
            ('Market Base Rent (UOM)', '$ / SF / Year'),
            ('Market Base Rent (New)', assumptions['market_rent_psf']),
            ('Market Base Rent (Renewal)', assumptions['market_rent_psf']),
            ('Market Base Rent (Blended)', assumptions['market_rent_psf']),
            ('', ''),
            ('Market Rental Value (UOM)', 'Continue Prior'),
            ('Market Rental Value', 'Continue Prior'),
            ('Use Market or Prior', 'N/A'),
            ('Prior Rent', 'N/A'),
            ('', ''),
            ('Rent Increases(UOM)', '% Increase'),
            ('Fixed Steps', f'{assumptions["market_escalation_rate"]:.2%}'),
            ('CPI Increase', 'None'),
            ('', ''),
            ('New Free Rent (Months)', 0),
            ('Renewal Free Rent (Months)', 0),
            ('Blended Free Rent (Months)', 0),
            ('', ''),
            ('Recovery Type', 'Continue Prior'),
            ('Miscellaneous Rent', 'None'),
            ('Incentives', 'None'),
            ('', ''),
            ('Tenant Improvements (UOM)', '$ / Area'),
            ('Tenant Improvements (New)', assumptions['tenant_improvements_psf']),
            ('Tenant Improvements (Renew)', 0),
            ('Tenant Improvements (Blended)', 0.75),
            ('', ''),
            ('Leasing Commissions (New UOM)', '% by Lease Year'),
            ('Leasing Commissions (New)', 'Varies'),
            ('Leasing Commissions (Renew UOM)', '% by Lease Year'),
            ('Leasing Commissions (Renew)', '0.00%'),
            ('Leasing Commissions (Blended)', ''),
            ('', ''),
            ('Upon Expiration', f'${assumptions["market_rent_psf"]:.2f} base')
        ]
        
        for label, value in params:
            if label:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                if label and not label.endswith('UOM)'):
                    ws[f'B{row}'].font = self.blue_font
            row += 1
            
        # Special note for leasing commissions
        lc_year1_pct = assumptions['leasing_commission_year1_pct'] * 100
        lc_subsequent_pct = assumptions['leasing_commission_subsequent_pct'] * 100
        ws['B35'] = f'{lc_year1_pct:.0f}% Year 1, {lc_subsequent_pct:.1f}% thereafter'
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25


def main():
    """Example usage"""
    
    # Property details
    property_data = {
        'property_name': '120 Valleywood Markham',
        'address': '120 Valleywood Drive',
        'purchase_price': 17800000,
        'property_type': 'Industrial'
    }
    
    # Current lease details from rent roll
    lease_data = {
        'tenant_name': 'Sentrex Health Solutions Inc.',
        'lease_start': '03/01/2022',
        'lease_end': '02/29/2032',
        'lease_term_years': 10,
        'current_annual_rent': 853608.91,  # $14.21/SF * 60,071 SF
        'area_sf': 60071,
        'escalation_rate': 0.03  # 3% annual
    }
    
    # Market assumptions for upon expiry
    assumptions = {
        'valuation_date': 'January, 2026',
        'discount_rate': 0.08,
        'resale_rate': 0.08,
        'leveraged_cf_rate': 0.08,
        'leveraged_resale_rate': 0.08,
        'discount_method': 'Annual',
        'hold_period_years': 10,
        'residual_sale_date': 'December, 2035',
        'period_to_cap': '12 Months After Sale',
        'exit_cap_rate': 0.065,
        'gross_up_noi': 'No',
        'selling_costs': 0.00,
        'renewal_probability': 0.85,  # 85% chance tenant renews
        'market_rent_psf': 17.50,
        'market_escalation_rate': 0.035,  # 3.5% annual escalations on market lease
        'market_term_years': 5,
        'vacancy_months': 8,  # If tenant doesn't renew
        'tenant_improvements_psf': 5,  # $5/SF TI allowance
        'leasing_commission_year1_pct': 0.08,  # 8% of year 1 NOI
        'leasing_commission_subsequent_pct': 0.035  # 3.5% of NOI thereafter
    }
    
    # Create underwriting
    underwriter = CREUnderwriter()
    wb = underwriter.create_underwriting(property_data, lease_data, assumptions)
    
    # Save to outputs directory
    import os
    output_dir = os.path.join(os.path.dirname(__file__), 'outputs')
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, 'cre_underwriting.xlsx')
    wb.save(output_path)
    print(f"Underwriting package created: {output_path}")

    return output_path


if __name__ == '__main__':
    main()
