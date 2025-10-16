"""
Semantic Document Parser using LLM
Extracts required underwriting inputs from any combination of documents (PDFs/Excel)
regardless of format or document type.
"""

import ollama
import os
import json
from typing import Dict, List, Optional, Union
import pdfplumber
import openpyxl
from datetime import datetime


class SemanticDocumentParser:
    """
    LLM-powered document parser that extracts required underwriting inputs.
    Processes PDFs (with table structure preservation) and Excel files (all sheets).
    Uses local Llama 3.2 via Ollama.
    """

    def __init__(self, model: str = "llama3.2:3b"):
        """
        Initialize semantic parser with local Ollama model.

        Args:
            model: Ollama model name (defaults to llama3.2:3b)
        """
        self.model = model
        # Test if Ollama is running and model is available
        try:
            ollama.list()
        except Exception as e:
            raise ValueError(f"Ollama not available. Make sure it's running: {e}")

    def extract_underwriting_inputs(self, documents_text: str, is_multi_tenant: Optional[bool] = None) -> Dict:
        """
        Extract all required underwriting inputs from combined document text.

        Args:
            documents_text: Combined text from all documents
            is_multi_tenant: Optional flag to specify if multi-tenant (auto-detected if None)

        Returns:
            Dictionary with all required inputs for underwriting
        """
        prompt = f"""Extract commercial real estate underwriting data from these documents.

REQUIRED FIELDS TO EXTRACT:
1. Property Address - full civic address
2. Purchase Price - in dollars
3. Tenant Name(s) - full legal name(s)
4. Area (SF) - rentable area in square feet
5. Current Net Rent ($/SF/Year) - base rent ONLY, from "Annual Rent/Area" or similar column
6. CAM Recovery ($/SF/Year) - common area maintenance charges
7. Tax Recovery ($/SF/Year) - property tax recovery
8. Insurance Recovery ($/SF/Year) - insurance recovery (use 0 if not found)
9. Lease Start Date - format: YYYY-MM-DD
10. Lease End Date - format: YYYY-MM-DD
11. Annual Escalation (%) - calculate from rent steps (e.g., 14.21→14.63→15.07 = 3%)

IMPORTANT EXTRACTION RULES:
- "Annual Rent/Area" or "Rent/Area" column = Current Net Rent ($/SF/Year)
- Calculate escalation from year-over-year rent increases in rent steps
- CAM/Tax recoveries may be in separate "Charge Schedules" or "Recoveries" section
- For multi-tenant: extract data for EACH tenant as separate entries
- Use exact values from tables, don't approximate

Return JSON:
{{
  "property_address": "address",
  "purchase_price": null or amount,
  "property_type": "Industrial" or inferred type,
  "tenants": [
    {{
      "tenant_name": "full name",
      "area_sf": number,
      "current_rent_psf": number (annual $/SF, base rent only),
      "cam_psf": number,
      "tax_psf": number,
      "insurance_psf": number or 0,
      "lease_start": "YYYY-MM-DD",
      "lease_end": "YYYY-MM-DD",
      "annual_escalation": number (as percentage, e.g., 3.0 for 3%)
    }}
  ]
}}

Documents:
{documents_text[:6000]}

Return ONLY valid JSON."""

        response = ollama.chat(
            model=self.model,
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0, "num_ctx": 8192}
        )

        json_text = response['message']['content'].strip()

        # Clean up JSON from markdown
        if '```json' in json_text:
            json_text = json_text.split('```json')[1].split('```')[0].strip()
        elif '```' in json_text:
            json_text = json_text.split('```')[1].split('```')[0].strip()

        # Remove commas from numbers (e.g., 60,071.00 -> 60071.00)
        import re
        json_text = re.sub(r'(\d+),(\d+)', r'\1\2', json_text)

        try:
            data = json.loads(json_text)
            return data
        except json.JSONDecodeError as e:
            print(f"Failed to parse JSON: {e}")
            print(f"Raw response: {json_text[:500]}")
            return {"error": "Failed to extract data", "raw_response": json_text}

    def extract_rent_roll(self, document_text: str, table_data: list = None) -> Dict:
        """
        Extract rent roll data for multi-tenant properties using structured table extraction.
        If table_data is provided, use it for LLM mapping; otherwise, fallback to text.
        """
        if table_data:
            # Pass the table as JSON to the LLM for mapping
            prompt = f"""
You are a commercial real estate rent roll parser. You are given a rent roll table as JSON (list of dicts, each dict is a row, keys are column headers). Your job is to:
1. For each unique tenant, identify the CURRENT lease term based on today's date. If all lease terms are in the future, use the one that starts earliest.
2. Use the rent from the CURRENT lease term for `current_rent_psf`.
3. List all future rent steps in a `rent_schedule` array.
4. Calculate `escalation_rate` as the average annual % increase in rent from the rent schedule.
5. Use the full tenant name, not a partial one.
6. Return this JSON structure:
1. Identify columns for: unit, tenant_name, area_sf, lease_start, lease_end.
2. Identify the column for annual rent per SF (often 'Annual Rent/Area') and use it for `current_rent_psf`.
3. Identify columns for CAM and Tax recoveries per SF and map them to `cam_psf` and `tax_psf`. If not present, use `null`.
4. If there are multiple rent steps for a single tenant, use the *current* active lease for the main tenant details. Calculate `escalation_rate` as the average annual percentage increase between the rent steps.
5. Use the full, complete tenant name.
6. Return ONLY the following JSON structure:
{{
    "property_address": "full property address",
    "total_area_sf": total rentable area in SF,
    "tenants": [
        {{
            "unit": "unit number",
            "tenant_name": "COMPLETE tenant name",
            "area_sf": area in square feet,
            "current_rent_psf": annual rent per SF (from Annual Rent/Area column),
            "lease_start": "MM/DD/YYYY",
            "lease_end": "MM/DD/YYYY",
            "renewal_option": "renewal terms if mentioned",
            "market_rent_psf": market rent if specified,
            "escalation_rate": calculated average annual escalation (e.g., 0.03 for 3%),
            "cam_psf": cam recovery per SF if present,
            "tax_psf": tax recovery per SF if present
        }}
    ]
}}

Table data:
{json.dumps(table_data)[:6000]}

IMPORTANT: Return ONLY valid JSON. Do NOT return Python code, explanations, or markdown. Only output the JSON object as specified above.
"""
            response = ollama.chat(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                options={"temperature": 0}
            )
            json_text = response['message']['content'].strip()
        else:
            # Fallback to text-based extraction (legacy)
            prompt = f"""
You are extracting data from a structured rent roll table.

IMPORTANT RULES:
1. "Annual Rent/Area" column = current_rent_psf ($/SF/year)
2. Calculate escalation_rate from rent steps (e.g., 14.21 → 14.63 = 3% escalation)
3. Use FULL tenant name, not partial
4. Area is in SF (square feet)

Extract tenant information and return this JSON structure:
{{
    "property_address": "full property address",
    "total_area_sf": total rentable area in SF,
    "tenants": [
        {{
            "unit": "unit number",
            "tenant_name": "COMPLETE tenant name",
            "area_sf": area in square feet,
            "current_rent_psf": annual rent per SF (from Annual Rent/Area column),
            "lease_start": "MM/DD/YYYY",
            "lease_end": "MM/DD/YYYY",
            "renewal_option": "renewal terms if mentioned",
            "market_rent_psf": market rent if specified,
            "escalation_rate": calculate from rent steps (e.g., 0.03 for 3%)
        }}
    ]
}}

Document:
{document_text[:4000]}

IMPORTANT: Return ONLY valid JSON. Do NOT return Python code, explanations, or markdown. Only output the JSON object as specified above.
"""
            response = ollama.chat(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                options={"temperature": 0}
            )
            json_text = response['message']['content'].strip()

        # Extract JSON from markdown code blocks if present
        if '```json' in json_text:
            json_text = json_text.split('```json')[1].split('```')[0].strip()
        elif '```' in json_text:
            json_text = json_text.split('```')[1].split('```')[0].strip()

        try:
            data = json.loads(json_text)
            return data
        except json.JSONDecodeError as e:
            print(f"Failed to parse JSON: {e}")
            print(f"Raw response: {json_text[:500]}")
            return {"error": "Failed to parse document", "raw_response": json_text}

    def extract_tax_bill(self, document_text: str) -> Dict:
        """
        Extract property tax information from tax bill.

        Returns structured tax data including assessment values, rates, and amounts.
        """
        prompt = f"""Extract property tax information from this tax bill.

Return a JSON object with this exact structure:
{{
  "property_address": "civic address from bill",
  "roll_number": "assessment roll number",
  "billing_date": "MM/DD/YYYY",
  "assessment_value": total assessed value (number),
  "assessment_breakdown": {{
    "commercial": assessed value for commercial portion,
    "industrial": assessed value for industrial portion,
    "other": other assessed values
  }},
  "tax_amounts": {{
    "municipal": municipal tax amount,
    "education": education tax amount,
    "total": total tax amount
  }},
  "due_dates": [
    {{"date": "MM/DD/YYYY", "amount": installment amount}}
  ]
}}

Document:
{document_text}

Return ONLY valid JSON, no explanations."""

        response = ollama.chat(
            model=self.model,
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0}
        )

        json_text = response['message']['content'].strip()

        # Clean up JSON
        if '```json' in json_text:
            json_text = json_text.split('```json')[1].split('```')[0].strip()
        elif '```' in json_text:
            json_text = json_text.split('```')[1].split('```')[0].strip()

        try:
            data = json.loads(json_text)
            return data
        except json.JSONDecodeError as e:
            print(f"Failed to parse JSON: {e}")
            return {"error": "Failed to parse tax bill", "raw_response": json_text}

    def extract_cam_expenses(self, document_text: str) -> Dict:
        """
        Extract CAM (Common Area Maintenance) expenses.

        Returns structured operating expense data by category.
        """
        prompt = f"""Extract ALL operating expense categories and amounts from this CAM expenses document.

Return a JSON object with this exact structure:
{{
  "year": tax/expense year,
  "total_area_sf": total area for expense allocation,
  "expenses": [
    {{
      "category": "expense category name",
      "amount": dollar amount (number),
      "per_sf": amount per square foot if calculable
    }}
  ],
  "subtotal": subtotal before admin fee,
  "admin_fee_pct": admin fee percentage (as decimal, e.g., 0.15 for 15%),
  "admin_fee_amount": admin fee dollar amount,
  "total": total expenses including admin fee,
  "total_per_sf": total per square foot
}}

Document:
{document_text}

Return ONLY valid JSON, no explanations."""

        response = ollama.chat(
            model=self.model,
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0}
        )

        json_text = response['message']['content'].strip()

        # Clean up JSON
        if '```json' in json_text:
            json_text = json_text.split('```json')[1].split('```')[0].strip()
        elif '```' in json_text:
            json_text = json_text.split('```')[1].split('```')[0].strip()

        try:
            data = json.loads(json_text)
            return data
        except json.JSONDecodeError as e:
            print(f"Failed to parse JSON: {e}")
            return {"error": "Failed to parse CAM expenses", "raw_response": json_text}

    def parse_file(self, file_path: str) -> Dict:
        """
        Parse any supported file type and extract structured data.
        For PDFs, extract tables as structured data for LLM mapping.
        """
        table_data = None
        if file_path.lower().endswith('.pdf'):
            document_text, table_data = self._extract_text_and_tables_from_pdf(file_path)
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            document_text = self._extract_text_from_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_path}")

        # Classify the document
        doc_type = self.classify_document(document_text)

        # Extract data based on type
        if doc_type == 'rent_roll':
            extracted_data = self.extract_rent_roll(document_text, table_data=table_data)
        elif doc_type == 'tax_bill':
            extracted_data = self.extract_tax_bill(document_text)
        elif doc_type == 'cam_expenses':
            extracted_data = self.extract_cam_expenses(document_text)
        else:
            extracted_data = {"error": "Unknown document type"}

        return {
            "document_type": doc_type,
            "file_path": file_path,
            "extracted_data": extracted_data
        }

    def _extract_text_and_tables_from_pdf(self, file_path: str):
        """
        Extract text and tables from PDF file.
        Returns (text, table_data) where table_data is a list of dicts (if tables found), else None.
        """
        text_parts = []
        all_tables = []

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Extract tables as structured data
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        # Convert table to list of dicts if header row is present
                        if len(table) > 1:
                            headers = [str(h).strip() if h else "" for h in table[0]]
                            for row in table[1:]:
                                row_dict = {headers[i]: (str(cell).strip() if cell else "") for i, cell in enumerate(row)}
                                all_tables.append(row_dict)
                        else:
                            # Table with no header row, just add as list
                            all_tables.append(table)
                # Also get regular text for context
                text = page.extract_text()
                if text:
                    text_parts.append(f"\n=== PAGE {page_num + 1} TEXT ===")
                    text_parts.append(text)

        text_content = "\n".join(text_parts)
        table_data = all_tables if all_tables else None
        return text_content, table_data

    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text and tables from PDF file with structured table preservation."""
        text_parts = []

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Extract tables first (structured data)
                tables = page.extract_tables()

                if tables:
                    text_parts.append(f"=== PAGE {page_num + 1} TABLES ===")
                    for table_num, table in enumerate(tables):
                        text_parts.append(f"\n--- Table {table_num + 1} ---")
                        # Convert table to text representation
                        for row in table:
                            if row:
                                clean_row = [str(cell).strip() if cell else "" for cell in row]
                                text_parts.append(" | ".join(clean_row))

                # Also get regular text for context
                text = page.extract_text()
                if text:
                    text_parts.append(f"\n=== PAGE {page_num + 1} TEXT ===")
                    text_parts.append(text)

        return "\n".join(text_parts)

    def _extract_text_from_excel(self, file_path: str) -> str:
        """Extract all text from Excel file, processing ALL sheets with structure preservation."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"\n=== SHEET: {sheet_name} ===")

            # Track if we have data
            has_data = False

            for row in ws.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        if isinstance(cell, datetime):
                            row_text.append(cell.strftime('%Y-%m-%d'))
                        else:
                            row_text.append(str(cell).strip())
                    else:
                        row_text.append("")

                # Only add non-empty rows
                if any(cell for cell in row_text):
                    text_parts.append(" | ".join(row_text))
                    has_data = True

            if not has_data:
                text_parts.append("(Empty sheet)")

        wb.close()
        return "\n".join(text_parts)


def extract_inputs_from_documents(file_paths: List[str], model: str = "llama3.2:3b") -> Dict:
    """
    Extract all required underwriting inputs from multiple documents.
    Handles any combination of PDFs and Excel files.

    Args:
        file_paths: List of file paths (PDFs and/or Excel files)
        model: Ollama model name

    Returns:
        Dictionary with all required underwriting inputs
    """
    parser = SemanticDocumentParser(model=model)

    # Combine all document text
    all_text_parts = []

    for file_path in file_paths:
        try:
            if file_path.lower().endswith('.pdf'):
                text = parser._extract_text_from_pdf(file_path)
                all_text_parts.append(f"\n{'='*60}\nFILE: {os.path.basename(file_path)}\n{'='*60}\n{text}")
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                text = parser._extract_text_from_excel(file_path)
                all_text_parts.append(f"\n{'='*60}\nFILE: {os.path.basename(file_path)}\n{'='*60}\n{text}")
            else:
                print(f"Skipping unsupported file: {file_path}")
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            continue

    if not all_text_parts:
        return {"error": "No documents could be processed"}

    combined_text = "\n".join(all_text_parts)

    # Extract all required inputs using LLM
    result = parser.extract_underwriting_inputs(combined_text)

    return result


def parse_multiple_documents(file_paths: List[str], model: str = "llama3.2:3b") -> Dict:
    """
    Parse multiple documents and combine the data.

    Args:
        file_paths: List of file paths to parse
        model: Ollama model name

    Returns:
        Combined data from all documents
    """
    parser = SemanticDocumentParser(model=model)

    results = {
        "rent_roll": None,
        "tax_bill": None,
        "cam_expenses": None,
        "unknown_docs": []
    }

    for file_path in file_paths:
        try:
            parsed = parser.parse_file(file_path)
            doc_type = parsed["document_type"]

            if doc_type == "rent_roll":
                results["rent_roll"] = parsed["extracted_data"]
            elif doc_type == "tax_bill":
                results["tax_bill"] = parsed["extracted_data"]
            elif doc_type == "cam_expenses":
                results["cam_expenses"] = parsed["extracted_data"]
            else:
                results["unknown_docs"].append({
                    "file": file_path,
                    "data": parsed
                })
        except Exception as e:
            results["unknown_docs"].append({
                "file": file_path,
                "error": str(e)
            })

    return results


# Example usage
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python semantic_parser.py <file_path1> [file_path2 ...] [--model MODEL_NAME]")
        print("\nExtracts underwriting inputs from any combination of PDF/Excel files")
        print("\nExamples:")
        print("  python semantic_parser.py rent_roll.pdf")
        print("  python semantic_parser.py rent_roll.pdf tax_bill.xlsx cam_expenses.pdf")
        print("  python semantic_parser.py data.xlsx --model llama3.2:11b")
        sys.exit(1)

    # Parse arguments
    file_paths = []
    model = "llama3.2:3b"

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == "--model" and i + 1 < len(sys.argv):
            model = sys.argv[i + 1]
            i += 2
        else:
            file_paths.append(sys.argv[i])
            i += 1

    if not file_paths:
        print("Error: No files specified")
        sys.exit(1)

    # Extract inputs from all documents
    result = extract_inputs_from_documents(file_paths, model=model)

    print(json.dumps(result, indent=2))
